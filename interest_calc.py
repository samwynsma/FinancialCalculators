import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

class InterestCalculator:
    def __init__(self, file_exists=False):
        self.starting_value = 0.0
        self.interest_rate = 0.0
        self.time = -1
        self.period_type = "year"
        self.will_add = "yes"
        self.added_investment = 0.0
        self.file_exists = file_exists
    
    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def parse_int(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int)):
            return int(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return int(raw_value)
    
    def create_gui(self):
        window = tk.Toplevel()
        window.title("Interest Generation Calculator")
        window.geometry("460x460")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Interest Generation Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Enter your investment details below and click Calculate",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))  

        inv_frame = tk.Frame(window)
        inv_frame.pack(padx=20, pady=8, fill = "x")

        tk.Label(inv_frame, text="Starting investment amount ($):", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.starting_inv_entry = tk.Entry(inv_frame, width=28)
        self.starting_inv_entry.grid(row=0, column=1, pady=6)

        tk.Label(inv_frame, text="Annual interest rate (%):", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.interest_rate_entry = tk.Entry(inv_frame, width=28)
        self.interest_rate_entry.grid(row=1, column=1, pady=6)

        self.period_type_var = tk.StringVar(value="year")
        tk.Label(inv_frame, text="Period Type:", anchor="w").grid(row=2, column=0, sticky="w", pady=6)
        tk.Radiobutton(inv_frame, text="Year", variable=self.period_type_var, value="year").grid(row=2, column=1, sticky="w")
        tk.Radiobutton(inv_frame, text="Month", variable=self.period_type_var, value="month").grid(row=2, column=1)

        tk.Label(inv_frame, text="Duration of investment:", anchor="w").grid(row=3, column=0, sticky="w", pady=6)
        self.time_var = tk.Entry(inv_frame, width=28)
        self.time_var.grid(row=3, column=1, pady=6)

        self.will_add_var = tk.StringVar(value="yes")
        tk.Label(inv_frame, text="Add investment each period?", anchor="w").grid(row=4, column=0, sticky="w", pady=6)
        tk.Radiobutton(inv_frame, text="Yes", variable=self.will_add_var, value="yes").grid(row=4, column=1, sticky="w")
        tk.Radiobutton(inv_frame, text="No", variable=self.will_add_var, value="no").grid(row=4, column=1)

        tk.Label(inv_frame, text="Additional Investment ($):", anchor="w").grid(row=5, column=0, sticky="w", pady=6)
        self.added_investment_var = tk.Entry(inv_frame, width=28)
        self.added_investment_var.grid(row=5, column=1, pady=6)

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Calculate", width=16, command=self.on_calculate).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)

        window.grab_set()
        window.mainloop()
    
    def on_calculate(self):
        try:
            starting_value = self.parse_float(self.starting_inv_entry.get())
            interest_rate = self.parse_float(self.interest_rate_entry.get())
            time_value = self.parse_int(self.time_var.get())
            if self.period_type_var.get() == "year":
                period_type = "year"
            else:
                period_type = "month"
            if self.will_add_var.get() == "yes":
                will_add = "yes"
                added_investment = self.parse_float(self.added_investment_var.get())
            else:
                will_add = "no"
        except ValueError:
            messagebox.showerror("Input error", "Please enter valid numeric values for all fields.")
            return
        
        if starting_value <= 0.0:
            messagebox.showerror("Input error", "Starting investment must be greater than 0. We must start with something")
            return
        
        if interest_rate <= 0.0:
            messagebox.showerror("Input error", "Interest rate must be greater than 0%. This is an investment, not a loan.")
            return
        elif(interest_rate > 12.0):
            proceed = messagebox.askyesno(
                "High Interest",
                "Are you sure that your interest is greater than 12 percent? That is an awfully high interest rate.",
            )
            if not proceed:
                return
        
        if(time_value <= 0):
            messagebox.showerror("Input error", "Time must be greater than 0.")
            return
        
        if(will_add == "yes"):
            if(added_investment <= 0.0):
                messagebox.showerror("Input error", "Added investment must be greater than 0.")
                return
            self.added_investment = added_investment
        else:
            self.added_investment = 0.0
        
        self.starting_value = starting_value
        self.interest_rate = interest_rate
        self.time = time_value
        self.period_type = period_type

        self.file_exists = self.generate_loan_documents()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("Investment document complete", "Investment calculation saved to InterestCalculation.xlsx.")
    
    def generate_loan_documents(self):
        money_by_period = []
        money_by_period.append(self.starting_value)

        if(self.period_type == "month"):
            interest = self.interest_rate / 12
            duration_of_period = "Month"
        else:
            interest = self.interest_rate
            duration_of_period = "Year"

        current_money = self.starting_value
        for i in range(self.time):
            current_money = round(current_money * (1 + interest / 100), 2) + self.added_investment
            money_by_period.append(current_money)
        

        workbook = None
        worksheet = None
        
        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook("InterestCalculation.xlsx")
            worksheet = workbook.create_sheet()
        
        worksheet.title = "Investment %d" % int(self.starting_value)

        worksheet["A1"] = "Interest calculation"
        worksheet["A2"] = "Starting amount: $%.2f" % (self.starting_value)
        worksheet["A3"] = "Interest rate: %.4f percent per %s" % (interest, duration_of_period)
        worksheet["A4"] = "Additional investment of $%.2f per %s" % (self.added_investment, duration_of_period)
        worksheet["A5"] = "Total time investigated: %d %ss" % (self.time, duration_of_period)
        worksheet["B2"] = duration_of_period
        worksheet["C2"] = "Money at %s" % (duration_of_period)
        money_format = "$#,##0.00"


        for i in range(self.time + 1):
            print("Money at the beginning of %s %d: %d" % (duration_of_period, i, money_by_period[i]))
            worksheet["B%d" % (i+3)] = i
            worksheet["C%d" % (i+3)] = money_by_period[i]
            worksheet["C%d" % (i+3)].number_format = money_format
            
        
        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2
        
        self.generate_graph(worksheet)
        workbook.save("InterestCalculation.xlsx")
        print("Finished creating financial documents.")
        return self.file_exists
    
    def generate_graph(self, worksheet):
        chart = BarChart()
        chart.title = "Investment Amount over Time"
        chart.type = "col"
        chart.style = 10
        chart.x_axis.title = self.period_type
        chart.y_axis.title = "Money in Account ($)"

        values = Reference(worksheet, min_col = 3, min_row = 2, max_row = self.time+3)
        categories = Reference(worksheet, min_col = 2, min_row = 3, max_row = self.time+3)
        
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4

        worksheet.add_chart(chart, "E5")
        return


def generate_interest(fileExists):
    calculator = InterestCalculator(fileExists)
    calculator.create_gui()
    return calculator.file_exists

