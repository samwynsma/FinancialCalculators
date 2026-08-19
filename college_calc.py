import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook

class CollegeSavingsCalculator:
    def __init__(self, file_exists=False):
        self.excel_file = "InterestCalculation.xlsx"
        self.starting_value = -1.0
        self.interest_rate = -1.0
        self.monthly_invest = -1.0
        self.goal = -1.0
        self.file_exists = file_exists

    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def create_gui(self):
        window = tk.Toplevel()
        window.title("College Savings Calculator")
        window.geometry("500x420")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="College Savings Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Enter your account information below, and see how it will grow in 18 years.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))

        inv_frame = tk.Frame(window)
        inv_frame.pack(padx=20, pady=8, fill = "x")

        tk.Label(inv_frame, text="Starting Value ($):", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.start_entry = tk.Entry(inv_frame, width=28)
        self.start_entry.grid(row=0, column=1, pady=6)

        tk.Label(inv_frame, text="Growth Rate (%):", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.growth_entry = tk.Entry(inv_frame, width=28)
        self.growth_entry.grid(row=1, column=1, pady=6)

        tk.Label(inv_frame, text="Monthly investment ($):", anchor="w").grid(row=2, column=0, sticky="w", pady=6)
        self.monthly_entry = tk.Entry(inv_frame, width=28)
        self.monthly_entry.grid(row=2, column=1, pady=6)

        tk.Label(inv_frame, text="Monetary Goal ($):", anchor="w").grid(row=3, column=0, sticky="w", pady=6)
        self.goal_entry = tk.Entry(inv_frame, width=28)
        self.goal_entry.grid(row=3, column=1, pady=6)

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Total Calc", width=16, command=self.find_total).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Goal Calc", width=16, command=self.find_goal).grid(row=0, column=1, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=2, padx=6)

        window.grab_set()
        window.mainloop()

    def simulate_vals(self, has_add = True, has_goal = True):
        try:
            start = self.parse_float(self.start_entry.get())
            growth = self.parse_float(self.growth_entry.get())
            monthly_add = self.parse_float(self.monthly_entry.get())
            end_goal = self.parse_float(self.goal_entry.get())
        except ValueError:
            messagebox.showerror("Input error", "Please enter valid numeric values for all fields.")
            return
        
        if start < 0.0:
            messagebox.showerror("Input error", "Starting savings cannot be negative")
            return
        
        if growth <= 0.0:
            messagebox.showerror("Input error", "Interest rate must be positive. Don't put your money in a no-growth account")
            return
        elif(growth > 12.0):
            proceed = messagebox.askyesno(
                "High Interest",
                "Are you sure that your interest is greater than 12 percent? That is an awfully high interest rate.",
            )
            if not proceed:
                return
        
        if monthly_add < 0.0 and has_add:
            messagebox.showerror("Input error", "Monthly additions to account cannot be negative")
            return
        
        if end_goal <= 0.0 and has_goal:
            messagebox.showerror("Input error", "End goal must be positive.")
            return
        
        self.starting_value = start
        self.interest_rate = growth
        self.monthly_invest = monthly_add
        self.goal = end_goal
        

    def find_total(self):
        self.simulate_vals(True, False)
        self.file_exists = self.generate_total_savings()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("College Savings Account Calc complete", "Retirement calculation saved to InterestCalculation.xlsx.")
        return
    
    def find_goal(self):
        self.simulate_vals(False, True)
        self.file_exists = self.generate_college_requirement()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("College Savings Account Calc complete", "Retirement calculation saved to InterestCalculation.xlsx.")
        return
    
    def generate_total_savings(self):
        months_to_graduate = 216
        current_money = self.starting_value
        growth = self.interest_rate / 12.0
        monthly = self.monthly_invest

        money = []
        money.append(current_money)

        for i in range(months_to_graduate):
            current_money = (current_money * (1 + growth / 100.0)) + monthly
            money.append(current_money)
        
        worksheet = None
        workbook = None

        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook(self.excel_file)
            worksheet = workbook.create_sheet()
        
        worksheet.title = "College Start %d" % int(self.starting_value)

        worksheet["A1"] = "College savings by end date."
        worksheet["A2"] = "Starting Money: $%.2f" % self.starting_value
        worksheet["A3"] = "Monthly addition: $%.2f" % self.monthly_invest
        worksheet["A4"] = "Growth rate: %.4f percent per month" % growth
        worksheet["B1"] = "Month"
        worksheet["C1"] = "Money in College Account"
        money_format = "$#,##0.00"

        for i in range(len(money)):
            worksheet["B%d" % (i+3)] = i
            worksheet["C%d" % (i+3)] = money[i]
            worksheet["C%d" % (i+3)].number_format = money_format
        
        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2
        

        workbook.save(self.excel_file)
        print("Finished creating college savings documents.")
        return self.file_exists
    
    def generate_college_requirement(self):
        months_to_graduate = 216
        current_money = self.starting_value
        growth = self.interest_rate / 12.0
        goal = self.goal

        money = []
        money.append(current_money)

        interest_part = (growth / 100.0) / (pow(1 + growth / 100.0, months_to_graduate) - 1)
        equity_part = goal - (current_money * pow(1 + growth / 100.0, months_to_graduate))
        monthly = interest_part * equity_part

        for i in range(months_to_graduate):
            current_money = (current_money * (1 + growth / 100.0)) + monthly
            money.append(current_money)
        
        worksheet = None
        workbook = None

        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook(self.excel_file)
            worksheet = workbook.create_sheet()
        
        worksheet.title = "College Goal %d" % int(self.goal)

        worksheet["A1"] = "College savings to get to goal."
        worksheet["A2"] = "Starting Money: $%.2f" % self.starting_value
        worksheet["A3"] = "Goal: $%.2f" % goal
        worksheet["A4"] = "Growth rate: %.4f percent per month" % growth
        worksheet["A5"] = "Money to save: %.2f per month" % monthly
        worksheet["B1"] = "Month"
        worksheet["C1"] = "Money in College Account"
        money_format = "$#,##0.00"

        for i in range(len(money)):
            worksheet["B%d" % (i+3)] = i
            worksheet["C%d" % (i+3)] = money[i]
            worksheet["C%d" % (i+3)].number_format = money_format
        
        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2
        

        workbook.save(self.excel_file)
        print("Finished creating college savings documents.")
        return self.file_exists



def college_save(file_exists):
    calculator = CollegeSavingsCalculator(file_exists)
    calculator.create_gui()
    return file_exists