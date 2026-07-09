import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook

class FoodCostCalculator:
    def __init__(self, file_exists):
        self.family_size = -1
        self.meals_per_day = -1
        self.snacks_per_day = -1
        self.current_food_spend = -1.0
        self.cost_per_meal = -1.0
        self.cost_per_snack = -1.0
        self.file_exists = file_exists
    
    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def parse_int(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int)):
            return int(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return int(raw_value)

    def create_gui(self):
        window = tk.Toplevel()
        window.title("Food Cost Calculator")
        window.geometry("460x460")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Food Cost Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Using the size of your family, the number of meals that you eat per person plus snacks, and the current food spend, I will determine how much you can save on a budget plan.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))

        food_frame = tk.Frame(window)
        food_frame.pack(padx=20, pady=8, fill="x")

        tk.Label(food_frame, text="Number of people in your family:", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.family_entry = tk.Entry(food_frame, width=28)
        self.family_entry.grid(row=0, column=1, pady=6)

        tk.Label(food_frame, text="Average meals per day per person:", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.day_meals_entry = tk.Entry(food_frame, width=28)
        self.day_meals_entry.grid(row=1, column=1, pady=6)

        tk.Label(food_frame, text="Snacks per day per person:", anchor="w").grid(row=2, column=0, sticky="w", pady=6)
        self.day_snacks_entry = tk.Entry(food_frame, width=28)
        self.day_snacks_entry.grid(row=2, column=1, pady=6)

        tk.Label(food_frame, text="Current money spent on food per week ($):", anchor="w").grid(row=3, column=0, sticky="w", pady=6)
        self.current_spend_entry = tk.Entry(food_frame, width=28)
        self.current_spend_entry.grid(row=3, column=1, pady=6)

        tk.Label(food_frame, text="Target Spend per meal ($):", anchor="w").grid(row=4, column=0, sticky="w", pady=6)
        self.target_meal_entry = tk.Entry(food_frame, width=28)
        self.target_meal_entry.grid(row=4, column=1, pady=6)

        tk.Label(food_frame, text="Target Spend per snack ($):", anchor="w").grid(row=5, column=0, sticky="w", pady=6)
        self.target_snack_entry = tk.Entry(food_frame, width=28)
        self.target_snack_entry.grid(row=5, column=1, pady=6)

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Calculate", width=16, command=self.on_calculate).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)

        window.grab_set()
        window.mainloop()

    def on_calculate(self):
        try:
            family = self.parse_int(self.family_entry.get())
            meals = self.parse_int(self.day_meals_entry.get())
            snacks = self.parse_int(self.day_snacks_entry.get())
            current_spend = self.parse_float(self.current_spend_entry.get())
            target_meal = self.parse_float(self.target_meal_entry.get())
            target_snack = self.parse_float(self.target_snack_entry.get())
        except:
            messagebox.showerror("Input error", "Please enter valid numeric values for all fields.")
            return
        
        if(family < 1):
            messagebox.showerror("Input error", "Family size must be positive.")
            return
        
        if(meals < 1):
            messagebox.showerror("Input error", "You must have at least one, preferably three meals a day.")
            return
        
        if(snacks < 0):
            messagebox.showerror("Input error", "Snacks cannot be negative.")
            return
        
        if(current_spend < 0.0):
            messagebox.showerror("Input error", "Amount spent on food cannot be negative.")
            return
        
        if(target_meal < 0.0):
            messagebox.showerror("Input error", "Target meal spend cannot be negative.")
            return
        
        if(target_snack < 0.0):
            messagebox.showerror("Input error", "Target snack spend cannot be negative.")
            return
        
        self.family_size = family
        self.meals_per_day = meals
        self.snacks_per_day = snacks
        self.current_food_spend = current_spend
        self.cost_per_meal = target_meal
        self.cost_per_snack = target_snack
        
        self.file_exists = self.generate_food_cost_doc()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("Inheritance document complete", "Inheritance saved to InterestCalculation.xlsx.")

    def generate_food_cost_doc(self):
        family = self.family_size
        day_meals = self.meals_per_day * family
        day_snacks = self.snacks_per_day * family
        day_meal_cost = self.cost_per_meal * day_meals
        day_snack_cost = self.cost_per_snack * day_snacks
        est_meal_cst = self.current_food_spend / (7 * day_meals + 3.5 * day_snacks)
        est_snack_cst = est_meal_cst / 2.0

        money_format = "$#,##0.00"

        cost_set = []
        cost_set_1 = [est_meal_cst, est_snack_cst]
        cost_set_2 = [day_meal_cost, day_snack_cost]
        cost_set_3 = [3, 1]
        cost_set_4 = [5, 2]
        cost_set_5 = [10, 4]
        cost_set_6 = [20, 8]
        cost_set.append(cost_set_1)
        cost_set.append(cost_set_2)
        cost_set.append(cost_set_3)
        cost_set.append(cost_set_4)
        cost_set.append(cost_set_5)
        cost_set.append(cost_set_6)

        worksheet = None
        workbook = None

        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook("InterestCalculation.xlsx")
            worksheet = workbook.create_sheet()

        worksheet.title = "Food Cost Calculator %d %d %d" % (day_meals, day_snacks, family)

        worksheet["A1"] = "Food Cost Calculator"
        worksheet["A2"] = "Family Size: %d" % family
        worksheet["A3"] = "Daily Meals Total: %d" % day_meals
        worksheet["A4"] = "Daily Snacks Total: %d" % day_snacks
        worksheet["C2"] = "Current Usage"
        worksheet["D2"] = "Target Usage"
        worksheet["E2"] = "Low Cost Usage"
        worksheet["F2"] = "Modest Cost Usage"
        worksheet["G2"] = "High Cost Usage"
        worksheet["H2"] = "Luxury Lifestyle Usage"
        worksheet["B3"] = "Total Cost per Month"
        worksheet["B4"] = "Total Cost per Week"
        worksheet["B5"] = "Total Cost per Day"
        worksheet["B6"] = "Est Cost per Meal"
        worksheet["B7"] = "Est Cost per Snack"
        worksheet["B8"] = "Total Yearly Cost"
        worksheet["B9"] = "Savings Compared To Current"

        data_columns = "CDEFGH"

        for i in range(len(data_columns)):
            worksheet["%c6" % data_columns[i]] = cost_set[i][0]
            worksheet["%c7" % data_columns[i]] = cost_set[i][1]
            worksheet["%c5" % data_columns[i]] = cost_set[i][0] * day_meals + cost_set[i][1] * day_snacks
            for j in range(3, 10):
                worksheet["%c%d" % (data_columns[i], j)].number_format = money_format


        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2
        

        workbook.save("InterestCalculation.xlsx")
        print("Finished creating food cost documents.")
        return self.file_exists
    


def food_coster(file_exists):
    calculator = FoodCostCalculator(file_exists)
    calculator.create_gui()
    return file_exists