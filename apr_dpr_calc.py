import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook

class AppreciateDepreciateCalculator:
    def __init__(self, file_exists = False):
        self.file_exists = file_exists
        self.initial_value = 0.0
        self.is_growing = False
        self.freq = "yearly"
        self.apr_dpr_rate = 0.0
        self.asset_name = ""
    
    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def create_gui(self):
        window = tk.Toplevel()
        self.window = window
        window.title("Appreciation/Depreciation Calculator")
        window.geometry("460x520")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Appreciation and Depreciation Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Type in an item name, the current value, whether or not it is growing in value, and the yearly rate of growth.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))

        apr_frame = tk.Frame(window)
        apr_frame.pack(padx=20, pady=8, fill="x")

        tk.Label(apr_frame, text="Asset Name:", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.item_entry = tk.Entry(apr_frame, width=28)
        self.item_entry.grid(row=0, column=1, pady=6)

        tk.Label(apr_frame, text="Current Value:", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.value_entry = tk.Entry(apr_frame, width=28)
        self.value_entry.grid(row=1, column=1, pady=6)

        self.pay_freq_var = tk.StringVar(value="increasing")
        tk.Label(apr_frame, text="Appreciating or Depreciating", anchor="w").grid(row=2, column=0, sticky="w", pady=6)
        inc_dec_frame = tk.Frame(apr_frame)
        inc_dec_frame.grid(row=2, column=1, columnspan=5, sticky="w", pady=6)
        tk.Radiobutton(inc_dec_frame, text="Apr", variable=self.pay_freq_var, value="increasing").pack(side="left", padx=(0, 12))
        tk.Radiobutton(inc_dec_frame, text="Dpr", variable=self.pay_freq_var, value="decreasing").pack(side="left", padx=(0, 12))

        tk.Label(apr_frame, text="Yearly Growth (%):", anchor="w").grid(row=3, column=0, sticky="w", pady=6)
        self.value_entry = tk.Entry(apr_frame, width=28)
        self.value_entry.grid(row=3, column=1, pady=6)

        self.period_var = tk.StringVar(value="yearly")
        tk.Label(apr_frame, text="Frequency of Calculation", anchor="w").grid(row=4, column=0, sticky="w", pady=6)
        period_frame = tk.Frame(apr_frame)
        period_frame.grid(row=4, column=1, columnspan=5, sticky="w", pady=6)
        tk.Radiobutton(period_frame, text="Year", variable=self.pay_freq_var, value="yearly").pack(side="left", padx=(0, 12))
        tk.Radiobutton(period_frame, text="Month", variable=self.pay_freq_var, value="monthly").pack(side="left", padx=(0, 12))

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Calculate", width=16, command=self.on_calculate).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)

        window.grab_set()
        window.mainloop()

    def on_calculate(self):
        try:
            initial = self.parse_float(self.item_entry.get())
            growth = self.parse_float(self.value_entry.get())
        except:
            messagebox.showerror("Input error", "Please enter valid numeric values for all fields.")
            return

        if initial <= 0.0:
            messagebox.showerror("Input error", "Initial value must be positive.")
            return

        if growth < 0.0:
            messagebox.showerror("Input error", "Growth must be positive.")
            return

        if growth == 0.0:
            messagebox.showerror("Input error", "If the growth is 0, the item is constant.")
            return

        if(self.pay_freq_var.get() == "decreasing"):
            growth = -growth

        self.initial_value = initial
        self.apr_dpr_rate = growth
        self.is_growing = self.pay_freq_var.get()
        self.asset_name = self.item_entry.get()
        self.freq = self.period_var.get()

        self.file_exists = self.generate_apr_document()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("Apr/Dpr document complete", "Apr/Dpr saved to InterestCalculation.xlsx.")

    
    def generate_apr_document(self):
        value_set = []

        worksheet = None
        workbook = None
        
        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook("InterestCalculation.xlsx")
            worksheet = workbook.create_sheet()

        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2
                
        
        workbook.save("InterestCalculation.xlsx")
        print("Finished creating apr/dpr documents.")
        return self.file_exists
    


def increase_decrease(file_exists):
    calculator = AppreciateDepreciateCalculator(file_exists)
    calculator.create_gui()
    return calculator.file_exists