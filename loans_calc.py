import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook


class LoanCalculator:
    def __init__(self, file_exists=False):
        self.starting_loan = 0.0
        self.interest_rate = 0.0
        self.amount_paid = 0.0
        self.file_exists = file_exists

    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)

    def create_gui(self):
        window = tk.Toplevel()
        window.title("Loan Payoff Calculator")
        window.geometry("460x360")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Loan Payoff Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Enter your loan details below and click Calculate.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))

        form_frame = tk.Frame(window)
        form_frame.pack(padx=20, pady=8, fill="x")

        tk.Label(form_frame, text="Starting loan amount ($):", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.starting_loan_entry = tk.Entry(form_frame, width=28)
        self.starting_loan_entry.grid(row=0, column=1, pady=6)

        tk.Label(form_frame, text="Annual interest rate (%):", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.interest_rate_entry = tk.Entry(form_frame, width=28)
        self.interest_rate_entry.grid(row=1, column=1, pady=6)

        tk.Label(form_frame, text="Monthly payment ($):", anchor="w").grid(row=2, column=0, sticky="w", pady=6)
        self.amount_paid_entry = tk.Entry(form_frame, width=28)
        self.amount_paid_entry.grid(row=2, column=1, pady=6)

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Calculate", width=16, command=self.on_calculate).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)

        window.grab_set()
        window.mainloop()

    def on_calculate(self):
        try:
            starting_loan = self.parse_float(self.starting_loan_entry.get())
            interest_rate = self.parse_float(self.interest_rate_entry.get())
            amount_paid = self.parse_float(self.amount_paid_entry.get())
        except ValueError:
            messagebox.showerror("Input error", "Please enter valid numeric values for all fields.")
            return

        if starting_loan <= 0.0:
            messagebox.showerror("Input error", "Starting loan must be greater than 0.")
            return

        if interest_rate <= 0.0:
            messagebox.showerror("Input error", "Interest rate must be greater than 0%.")
            return

        minimum_payment = starting_loan * (interest_rate / 1200.0)
        if amount_paid <= minimum_payment:
            messagebox.showerror(
                "Input error",
                "Monthly payment must be greater than the monthly interest amount: $%.2f." % minimum_payment,
            )
            return

        self.starting_loan = starting_loan
        self.interest_rate = interest_rate
        self.amount_paid = amount_paid

        self.file_exists = self.generate_loan_documents()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("Loan payoff complete", "Loan payoff calculation saved to InterestCalculation.xlsx.")

    def generate_loan_documents(self):
        loan_remaining = [self.starting_loan]
        interest = self.interest_rate / 12.0
        current_money = self.starting_loan

        while current_money > 0.0 and len(loan_remaining) <= 600:
            current_money = round(current_money * (1 + interest / 100.0), 2) - self.amount_paid
            loan_remaining.append(max(current_money, 0.0))

        if not self.file_exists:
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook("InterestCalculation.xlsx")
            worksheet = workbook.create_sheet()

        worksheet.title = "Loan %d" % int(self.starting_loan)
        worksheet["A1"] = "Loan calculation"
        worksheet["A2"] = "Starting amount: $%.2f" % self.starting_loan
        worksheet["A3"] = "Interest rate: %.4f percent per month" % interest
        worksheet["A4"] = "Loan payment of $%.2f per month" % self.amount_paid

        if len(loan_remaining) < 600:
            years = (len(loan_remaining) - 1) // 12
            months = (len(loan_remaining) - 1) % 12
            if years > 1 and months > 1:
                worksheet["A5"] = "Loan will be paid off in %d years and %d months" % (years, months)
            elif years == 1 and months > 1:
                worksheet["A5"] = "Loan will be paid off in 1 year and %d months" % months
            elif years > 1 and months == 1:
                worksheet["A5"] = "Loan will be paid off in %d years and 1 month" % years
            elif years == 1 and months == 1:
                worksheet["A5"] = "Loan will be paid off in 1 year and 1 month"
            elif months > 1:
                worksheet["A5"] = "Loan will be paid off in %d months" % months
            elif years > 1:
                worksheet["A5"] = "Loan will be paid off in %d years" % years
            elif years == 1:
                worksheet["A5"] = "Loan will be paid off in 1 year"
            else:
                worksheet["A5"] = "Loan will be paid off in 1 month"
        else:
            worksheet["A5"] = "Loan will take longer than 50 years to pay off."

        if len(loan_remaining) < 600:
            total_amount = self.amount_paid * (len(loan_remaining) - 2) + (loan_remaining[-2] * (1 + interest / 100.0))
            worksheet["A6"] = "Total amount of money paid is %.2f" % total_amount

        worksheet["B2"] = "Month"
        worksheet["C2"] = "Remaining loan"
        money_format = "$#,##0.00"

        for i, balance in enumerate(loan_remaining):
            worksheet["B%d" % (i + 3)] = i
            worksheet["C%d" % (i + 3)] = balance
            worksheet["C%d" % (i + 3)].number_format = money_format

        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > length:
                        length = len(str(cell.value))
                except Exception:
                    pass
            worksheet.column_dimensions[column].width = length + 2

        workbook.save("InterestCalculation.xlsx")
        return self.file_exists


def loan_payoff(fileExists):
    calculator = LoanCalculator(fileExists)
    calculator.create_gui()
    return calculator.file_exists
