import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook

class TakeHomeCalculator:
    def __init__(self, file_exists = False):
        self.pay_rate = -1.0
        self.percent_tax_rate = -1.0
        self.set_aside = -1.0
        self.pay_freq = "hourly"
        self.period = "monthly"
        self.file_exists = file_exists

    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def create_gui(self):
        window = tk.Toplevel()
        window.title("Take Home Pay Calculator")
        window.geometry("500x420")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Take Home Pay Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Enter your pay and tax information below to see what your take home will be.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))

        inv_frame = tk.Frame(window)
        inv_frame.pack(padx=20, pady=8, fill = "x")

        tk.Label(inv_frame, text="Pay Rate ($):", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.pay_entry = tk.Entry(inv_frame, width=28)
        self.pay_entry.grid(row=0, column=1, pady=6)

        self.pay_freq_var = tk.StringVar(value="hourly")
        tk.Label(inv_frame, text="Pay Frequency:", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        pay_freq_frame = tk.Frame(inv_frame)
        pay_freq_frame.grid(row=1, column=1, columnspan=5, sticky="w", pady=6)
        tk.Radiobutton(pay_freq_frame, text="Year", variable=self.pay_freq_var, value="yearly").pack(side="left", padx=(0, 12))
        tk.Radiobutton(pay_freq_frame, text="Month", variable=self.pay_freq_var, value="monthly").pack(side="left", padx=(0, 12))
        tk.Radiobutton(pay_freq_frame, text="Week", variable=self.pay_freq_var, value="weekly").pack(side="left", padx=(0, 12))
        tk.Radiobutton(pay_freq_frame, text="Day", variable=self.pay_freq_var, value="daily").pack(side="left", padx=(0, 12))
        tk.Radiobutton(pay_freq_frame, text="Hour", variable=self.pay_freq_var, value="hourly").pack(side="left")

        tk.Label(inv_frame, text="Tax Rate (%):", anchor="w").grid(row=2, column=0, sticky="w", pady=6)
        self.tax_entry = tk.Entry(inv_frame, width=28)
        self.tax_entry.grid(row=2, column=1, pady=6)

        tk.Label(inv_frame, text="Set Aside (%):", anchor="w").grid(row=3, column=0, sticky="w", pady=6)
        self.aside_entry = tk.Entry(inv_frame, width=28)
        self.aside_entry.grid(row=3, column=1, pady=6)

        self.take_home_freq_var = tk.StringVar(value = "monthly")
        tk.Label(inv_frame, text="Take Home Frequency:", anchor="w").grid(row=4, column=0, sticky="w", pady=6)
        take_home_freq_frame = tk.Frame(inv_frame)
        take_home_freq_frame.grid(row=4, column=1, columnspan=5, sticky="w", pady=6)
        tk.Radiobutton(take_home_freq_frame, text="Year", variable=self.take_home_freq_var, value="yearly").pack(side="left", padx=(0, 12))
        tk.Radiobutton(take_home_freq_frame, text="Month", variable=self.take_home_freq_var, value="monthly").pack(side="left", padx=(0, 12))
        tk.Radiobutton(take_home_freq_frame, text="Week", variable=self.take_home_freq_var, value="weekly").pack(side="left", padx=(0, 12))
        tk.Radiobutton(take_home_freq_frame, text="Day", variable=self.take_home_freq_var, value="daily").pack(side="left", padx=(0, 12))
        tk.Radiobutton(take_home_freq_frame, text="Hour", variable=self.take_home_freq_var, value="hourly").pack(side="left")

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Calculate", width=16, command=self.on_calculate).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)

        window.grab_set()
        window.mainloop()
    
    def on_calculate(self):
        try:
            pay = self.parse_float(self.pay_entry.get())
            tax = self.parse_float(self.tax_entry.get())
            set_aside = self.parse_float(self.aside_entry.get())
            freq_pay = self.pay_freq_var.get().lower()
            freq_take = self.take_home_freq_var.get().lower()
        except ValueError:
            messagebox.showerror("Input error", "Please enter valid numeric values for all fields.")
            return
        
        if(pay <= 0.0):
            messagebox.showerror("Input error", "Pay must be positive.")
            return
        
        if(tax < 0.0):
            messagebox.showerror("Input error", "Sadly, we don't live in a libertarian society.")
            return
        
        if(set_aside < 0.0):
            messagebox.showerror("Input error", "Set aside cannot be negative")
            return
        
        self.pay_rate = pay
        self.percent_tax_rate = tax
        self.set_aside = set_aside
        self.pay_freq = freq_pay
        self.period = freq_take

        self.file_exists = self.generate_take_home_document()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("Take home document", "Take home calculation saved to InterestCalculation.xlsx.")
    
    def generate_take_home_document(self):
        starting_pay = self.pay_rate
        starting_pay_freq = self.pay_freq
        ending_pay = 0
        ending_pay_freq = self.period
        ending_formula = []

        if starting_pay_freq == ending_pay_freq:
            ending_pay = starting_pay
        elif starting_pay_freq == "hourly":
            if ending_pay_freq == "daily":
                ending_pay = starting_pay * 8
            elif ending_pay_freq == "weekly":
                ending_pay = starting_pay * 40
            elif ending_pay_freq == "monthly":
                ending_pay = starting_pay * 160
            else:
                ending_pay = starting_pay * 2080
        elif starting_pay_freq == "daily":
            if ending_pay_freq == "hourly":
                ending_pay = starting_pay / 8
            elif ending_pay_freq == "weekly":
                ending_pay = starting_pay * 5
            elif ending_pay_freq == "monthly":
                ending_pay = starting_pay * 20
            else:
                ending_pay = starting_pay * 260
        elif starting_pay_freq == "weekly":
            if ending_pay_freq == "hourly":
                ending_pay = starting_pay / 40
            elif ending_pay_freq == "daily":
                ending_pay = starting_pay / 5
            elif ending_pay_freq == "monthly":
                ending_pay = starting_pay * 4
            else:
                ending_pay = starting_pay * 52
        elif starting_pay_freq == "monthly":
            if ending_pay_freq == "hourly":
                ending_pay = starting_pay / 160
            elif ending_pay_freq == "daily":
                ending_pay = starting_pay / 20
            elif ending_pay_freq == "weekly":
                ending_pay = starting_pay / 4
            else:
                ending_pay = starting_pay * 13
        else:
            if ending_pay_freq == "hourly":
                ending_pay = starting_pay / 2080
            elif ending_pay_freq == "daily":
                ending_pay = starting_pay / 260
            elif ending_pay_freq == "weekly":
                ending_pay = starting_pay / 52
            else:
                ending_pay = starting_pay / 13
        
        ending_formula.append(ending_pay)

        tax_take = (self.percent_tax_rate / 100.0) * ending_pay
        set_aside_take = (self.set_aside / 100.0) * ending_pay
        ending_pay_res = (ending_pay - tax_take - set_aside_take)

        worksheet = None
        workbook = None

        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook("InterestCalculation.xlsx")
            worksheet = workbook.create_sheet()
        
        worksheet.title = "Take Home Calc %d" % int(starting_pay)

        money_format = "$#,##0.00"

        worksheet["A1"] = "Take Home Calculator"
        worksheet["A2"] = "Base pay: $%.2f" % starting_pay
        worksheet["A3"] = "Pay frequency: %s" % starting_pay_freq
        worksheet["A4"] = "Goal pay frequency: %s" % ending_pay_freq
        worksheet["C2"] = "Details"
        worksheet["B3"] = "Adjusted Base Pay"
        worksheet["B4"] = "Tax Reduction"
        worksheet["B5"] = "Set Aside for 401K/IRA"
        worksheet["B6"] = "Total Reductions"
        worksheet["B7"] = "Amount Remaining"
        worksheet["C3"] = ending_pay
        worksheet["C3"].number_format = money_format
        worksheet["C4"] = tax_take
        worksheet["C4"].number_format = money_format
        worksheet["C5"] = set_aside_take
        worksheet["C5"].number_format = money_format
        worksheet["C6"] = tax_take + set_aside_take
        worksheet["C6"].number_format = money_format
        worksheet["C7"] = ending_pay_res
        worksheet["C7"].number_format = money_format

        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2

        workbook.save("InterestCalculation.xlsx")
        print("Finished creating take home pay documents.")
        return self.file_exists

    

def take_home(file_exists):
    calculator = TakeHomeCalculator(file_exists)
    calculator.create_gui()
    return calculator.file_exists