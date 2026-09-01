import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook

from excel_doc import ExcelDocument


class InheritanceCalculator:
    def __init__(self, file_exists=False, document=None):
        self.document = document if document is not None else ExcelDocument()
        self.excel_file = "InterestCalculation.xlsx"
        self.total_inheritance = 0.0
        self.estate_tax = -1.0
        self.inherit_percent = -1.0
        self.total_debt = -1.0
        self.personal_spend_percent = -1.0
        self.file_exists = file_exists

    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def create_gui(self):
        window = tk.Toplevel()
        window.title("Inheritance Calculator")
        window.geometry("460x460")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Inheritance Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Determine the Amount of Money that you will Receive from Inheritance.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))

        inh_frame = tk.Frame(window)
        inh_frame.pack(padx=20, pady=8, fill = "x")

        tk.Label(inh_frame, text="Total Inheritance Amount ($):", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.total_inh_entry = tk.Entry(inh_frame, width=28)
        self.total_inh_entry.grid(row=0, column=1, pady=6)

        tk.Label(inh_frame, text="Estate Tax (%):", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.tax_entry = tk.Entry(inh_frame, width=28)
        self.tax_entry.grid(row=1, column=1, pady=6)

        tk.Label(inh_frame, text="Percent Inherited (%):", anchor="w").grid(row=2, column=0, sticky="w", pady=6)
        self.percent_inherit_entry = tk.Entry(inh_frame, width=28)
        self.percent_inherit_entry.grid(row=2, column=1, pady=6)

        tk.Label(inh_frame, text="Total Debt ($):", anchor="w").grid(row=3, column=0, sticky="w", pady=6)
        self.debt_entry = tk.Entry(inh_frame, width=28)
        self.debt_entry.grid(row=3, column=1, pady=6)

        tk.Label(inh_frame, text="Personal Spend (%):", anchor="w").grid(row=4, column=0, sticky="w", pady=6)
        self.personal_entry = tk.Entry(inh_frame, width=28)
        self.personal_entry.grid(row=4, column=1, pady=6)

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Calculate", width=16, command=self.on_calculate).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)

        window.grab_set()
        window.mainloop()
    
    def on_calculate(self):
        try:
            inherit = self.parse_float(self.total_inh_entry.get())
            tax = self.parse_float(self.tax_entry.get())
            percent = self.parse_float(self.percent_inherit_entry.get())
            debt = self.parse_float(self.debt_entry.get())
            personal = self.parse_float(self.personal_entry.get())
        except ValueError:
            messagebox.showerror("Input error", "Please enter valid numeric values for all fields.")
            return
        
        if(inherit <= 0.0):
            messagebox.showerror("Input error", "Inheritance must be positive")
            return
        
        if(tax < 0.0):
            messagebox.showerror("Input error", "Tax cannot be negative")
            return
        if(tax > 100.0):
            messagebox.showerror("Input error", "Tax cannot be greater than 100 percent.")
            return

        if(percent <= 0.0):
            messagebox.showerror("Input error", "Must receive a portion of the inheritance to be relevant.")
            return
        
        if(debt < 0.0):
            messagebox.showerror("Input error", "Debt cannot be negative")
            return
        
        if(personal < 0.0):
            messagebox.showerror("Input error", "Personal fun spend cannot be negative")
            return
        if(personal > 100.0):
            messagebox.showerror("Personal spend cannot be greater than 100 percent.")
            return
        
        self.total_inheritance = inherit
        self.estate_tax = tax
        self.inherit_percent = percent
        self.total_debt = debt
        self.personal_spend_percent = personal

        self.file_exists = self.generate_inherit_document()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("Inheritance document complete", "Inheritance saved to InterestCalculation.xlsx.")
    
    def generate_inherit_document(self):
        inherit = self.total_inheritance
        tax = self.estate_tax
        percent = self.inherit_percent
        debt = self.total_debt
        personal = self.personal_spend_percent
        money_format = "$#,##0.00"

        total_inherited = (inherit * (1.0 - (tax / 100.0)) * (percent / 100.0))
        total_personal_spend = (total_inherited * (personal / 100.0))

        debt_percent = debt / total_inherited * 100.0

        worksheet = None
        workbook = None

        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook(self.excel_file)
            worksheet = workbook.create_sheet()

        worksheet.title = "Inheritance Calculator %d" % int(inherit)

        worksheet["A1"] = "Inheritance Calculation"
        worksheet["A2"] = "Initial Inheritance: $%.2f" % inherit
        worksheet["A3"] = "Estate Tax: %.2f" % tax
        worksheet["A4"] = "Percent Inherited: %.2f" % percent
        worksheet["A5"] = "Debt: $%.2f" % debt
        worksheet["A6"] = "Desired Personal Spend: %.2f" % personal
        worksheet["B3"] = "Total Inherited" 
        worksheet["B4"] = "Debt to pay off"
        worksheet["B5"] = "Personal Usage"
        worksheet["B6"] = "Remainder" 
        worksheet["C2"] = "Planned Usage"
        worksheet["C3"] = total_inherited
        worksheet["C4"] = debt
        worksheet["C5"] = min(total_personal_spend, total_inherited - debt)
        worksheet["C6"] = max(0, total_inherited - debt - min(total_personal_spend, total_inherited - debt))
        worksheet["D2"] = "Potential Amounts"
        worksheet["E2"] = "Personal Spend"
        worksheet["F2"] = "Amount Invested"

        personal_percent = 0
        while personal_percent < (100.0 - debt_percent):
            worksheet["D%d" % (personal_percent + 3)] = personal_percent
            worksheet["E%d" % (personal_percent + 3)] = total_inherited * (personal_percent / 100.0)
            worksheet["F%d" % (personal_percent + 3)] = total_inherited - debt - (total_inherited * (personal_percent / 100.0))
            worksheet["E%d" % (personal_percent + 3)].number_format = money_format
            worksheet["F%d" % (personal_percent + 3)].number_format = money_format
            personal_percent += 1
        
        for i in range(2, 7):
            worksheet["C%d" % i].number_format = money_format
        
        

        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2
        

        workbook.save(self.excel_file)
        print("Finished creating inheritance documents.")
        return self.file_exists


def inherit_money(document_or_file_exists=None):
    if isinstance(document_or_file_exists, ExcelDocument):
        calculator = InheritanceCalculator(document=document_or_file_exists)
    else:
        calculator = InheritanceCalculator(file_exists=document_or_file_exists)
    calculator.create_gui()
    return calculator.file_exists