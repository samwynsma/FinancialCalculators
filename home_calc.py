import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook

class HomeAffordabilityCalculator:
    def __init__(self, file_exists=False):
        self.down_payment = -1.0
        self.monthly_salary = -1.0
        self.interest_rate = -1.0
        self.additional_debt = -1.0
        self.mortgage_type = "30 year"
        self.file_exists = False

    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def create_gui(self):
        window = tk.Toplevel()
        window.title("Home Affordability Calculator")
        window.geometry("580x400")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Home Affordability Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Enter your monetary information below and see what kind of house you can afford. If you are lost on budgetary information, use the budget calculator to determine stats there.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))

        home_frame = tk.Frame(window)
        home_frame.pack(padx=20, pady=8, fill="x")

        tk.Label(home_frame, text="Down Payment ($):", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.down_payment_entry = tk.Entry(home_frame, width=28)
        self.down_payment_entry.grid(row=0, column=1, pady=6)

        tk.Label(home_frame, text="Monthly Post-Tax Salary ($):", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.post_tax_entry = tk.Entry(home_frame, width=28)
        self.post_tax_entry.grid(row=1, column=1, pady=6)

        tk.Label(home_frame, text="Additional Debt ($):", anchor="w").grid(row=2, column=0, sticky="w", pady=6)
        self.added_debt_entry = tk.Entry(home_frame, width=28)
        self.added_debt_entry.grid(row=2, column=1, pady=6)

        tk.Label(home_frame, text="Interest Rate (%):", anchor="w").grid(row=3, column=0, sticky="w", pady=6)
        self.interest_entry = tk.Entry(home_frame, width=28)
        self.interest_entry.grid(row=3, column=1, pady=6)

        self.loan_length_var = tk.StringVar(value = "30 year")
        tk.Label(home_frame, text="Loan Length:", anchor="w").grid(row=4, column=0, sticky="w", pady=6)
        loan_length_frame = tk.Frame(home_frame)
        loan_length_frame.grid(row=4, column=1, columnspan=5, sticky="w", pady=6)
        tk.Radiobutton(loan_length_frame, text="30 year", variable=self.loan_length_var, value="30 year").pack(side="left", padx=(0, 12))
        tk.Radiobutton(loan_length_frame, text="25 year", variable=self.loan_length_var, value="25 year").pack(side="left", padx=(0, 12))
        tk.Radiobutton(loan_length_frame, text="20 year", variable=self.loan_length_var, value="20 year").pack(side="left", padx=(0, 12))
        tk.Radiobutton(loan_length_frame, text="15 year", variable=self.loan_length_var, value="15 year").pack(side="left", padx=(0, 12))
        tk.Radiobutton(loan_length_frame, text="10 year", variable=self.loan_length_var, value="10 year").pack(side="left")

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))
        
        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Calculate", widt=16, command=self.on_calculate).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)

        window.grab_set()
        window.mainloop()
    
    def on_calculate(self):
        try:
            down_payment = self.parse_float(self.down_payment_entry.get())
            monthly = self.parse_float(self.post_tax_entry.get())
            debt = self.parse_float(self.added_debt_entry.get())
            interest = self.parse_float(self.interest_entry.get())
        except:
            messagebox.showerror("Input error", "Please enter valid numeric values for all fields.")
            return
        
        if(down_payment < 0.0):
            messagebox.showerror("Input error", "Down payment cannot be negative")
            return
        
        if(monthly <= 0.0 or monthly <= debt):
            messagebox.showerror("Input error", "Monthly post tax money must be greater than zero, and must be greater than current debt.")
            return
        
        if(debt < 0.0):
            messagebox.showerror("Input error", "Debt cannot be negative. If you have additional money coming in, put it in post tax.")
            return
        elif(debt * 2 >= monthly):
            proceed = messagebox.askyesno(
                "High Debt Warning",
                "At your debt level, you won't be able to afford a house that costs more than your down payment. Are you certain you want to proceed?"
            )
            if not proceed:
                return
        
        if(interest <= 0.0):
            messagebox.showerror("Input error", "Interest must be positive.")
            return
        
        self.down_payment = down_payment
        self.monthly_salary = monthly
        self.additional_debt = debt
        self.interest_rate = interest
        self.mortgage_type = self.loan_length_var.get()
        self.file_exists = self.generate_house_affordability()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("Home Affordability document", "Home Affordability calculation saved to InterestCalculation.xlsx.")
    
    def generate_house_affordability(self):
        salary = self.monthly_salary
        debt = self.additional_debt
        interest = self.interest_rate / 12.0
        down_payment = self.down_payment
        loan_type = self.mortgage_type
        monthly_payments = []
        mortgages = []
        percents = []
        months = 0
        if loan_type == "30 year":
            months = 360
        elif loan_type == "25 year":
            months = 300
        elif loan_type == "20 year":
            months = 240
        elif loan_type == "15 year":
            months = 180
        else:
            months = 120

        for i in range(10, 51):
            percents.append(i)
            month_pay = max(salary * (i / 100.0) - debt, 0)
            monthly_payments.append(month_pay)
            interest_numerator = (interest / 100.0) * pow((1 + (interest / 100.0)), months)
            interest_denom = pow(1 + (interest / 100.0), months) - 1
            total_loan = month_pay * interest_denom / interest_numerator
            mortgages.append(total_loan + down_payment)

        worksheet = None
        workbook = None

        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook("InterestCalculation.xlsx")
            worksheet = workbook.create_sheet()

        worksheet.title = "Home Calculator %d" % int(salary)

        worksheet["A1"] = "House Price by Percent"
        worksheet["A2"] = "Monthly Take Home: $%.2f" % salary
        worksheet["A3"] = "Existing Debt: $%.2f" % debt
        worksheet["A4"] = "Interest rate: %.4f percent per month" % interest
        worksheet["A5"] = "Down Payment: $%.2f" % down_payment
        worksheet["A6"] = "Loan Type: %s" % loan_type
        worksheet["B2"] = "Debt Percentage"
        worksheet["C2"] = "House Price"
        worksheet["D2"] = "Monthly Mortgage Payment"
        worksheet["E1"] = "Note: debt percentage greater than 50 percent is discouraged."
        worksheet["E2"] = "Most underwriters will not write someone in that case."
        money_format = "$#,##0.00"

        for i in range(len(percents)):
            worksheet["B%d" % (i+3)] = percents[i]
            worksheet["C%d" % (i+3)] = mortgages[i]
            worksheet["C%d" % (i+3)].number_format = money_format
            worksheet["D%d" % (i+3)] = monthly_payments[i]
            worksheet["D%d" % (i+3)].number_format = money_format

        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2
        

        workbook.save("InterestCalculation.xlsx")
        print("Finished creating home affordability documents.")
        return self.file_exists


def house_affordability(file_exists):
    calculator = HomeAffordabilityCalculator(file_exists)
    calculator.create_gui()
    return calculator.file_exists