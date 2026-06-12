import tkinter as tk
from tkinter import messagebox

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
import openpyxl

class BudgetMaker:
    def __init__(self, file_exists=False):
        self.salary = -1.0
        self.family_size = -1
        self.pets = -1
        self.cars = -1
        self.mortgage_rent = -1.0
        self.car_loan = -1.0
        self.other_debt = -1.0
        self.utilities = -1.0
        self.insurance = -1.0
        self.food = -1.0
        self.copays = -1.0
        self.giving = -1.0
        self.investments = -1.0
        self.extra = -1.0
        self.file_exists = file_exists
    
    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def parse_int(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int)):
            return int(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return int(raw_value)
    
    def create_gui(self):
        window = tk.Toplevel()
        window.title("Budget Calculator")
        window.geometry("460x460")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Budget Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Enter your budget details below and click Calculate. Enter a \"0\" into boxes that aren't relevant to your budget.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))  

        inv_frame = tk.Frame(window)
        inv_frame.pack(padx=20, pady=8, fill = "x")

        tk.Label(inv_frame, text="Monthly post-tax take home pay ($):", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.starting_inv_entry = tk.Entry(inv_frame, width=28)
        self.starting_inv_entry.grid(row=0, column=1, pady=6)

        tk.Label(inv_frame, text="Number of people in Family:", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.interest_rate_entry = tk.Entry(inv_frame, width=28)
        self.interest_rate_entry.grid(row=1, column=1, pady=6)

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Calculate", width=16, command=self.on_calculate).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)

        window.grab_set()
        window.mainloop()
    
    def on_calculate(self):
        return

def budget_maker(file_exists):
    calculator = BudgetMaker(file_exists)
    calculator.create_gui()
    return calculator.file_exists


#def budget_maker(fileExists):
    print("Welcome to the budget calculator. I'm going to ask you some questions about your life to determine your budget.")
    print("Furthermore, I'm going to ask what your current spending is, so we can do a side by side comparison.")
    print("At the end, we will see how much leftover money there is, and I will give some recommendations.")
    budget_types = ["monthly pay", "people supported", "pets supported", "cars owned", "mortgage or rent", "auto payments", "debt payments", "utilities", "insurance", "food", "medical", "giving", "investments", "extra"]
    budget_information = [-1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1]


    while budget_information[0] <= 0.0:
        num_input = input("How much money do you take home each month? This is after taxes, 401k, etc. ")
        try:
            budget_information[0] = int(num_input)
            if(budget_information[0] <= 0):
                print("I'm not doing this if you don't take home any monthly pay.")
        except:
            print("Invalid input: try putting a number there")

    while budget_information[1] <= 0:
        num_input = input("How many people are in your family? ")
        try:
            budget_information[1] = int(num_input)
            if budget_information[1] <= 0:
                print("Invalid input: you count as one of those people.")
        except:
            print("Invalid input: try putting a number there")
    
    while budget_information[2] < 0:
        num_input = input("How many pets do you have? ")
        try:
            budget_information[2] = int(num_input)
            if budget_information[2] < 0:
                print("Invalid input: you cannot have a negative number of pets.")
        except:
            print("Invalid input: try putting a number there")
    
    while budget_information[3] < 0:
        num_input = input("How many cars do you own? ")
        try:
            budget_information[3] = int(num_input)
            if budget_information[3] < 0:
                print("Invalid input: you cannot have a negative number of cars.")
        except:
            print("Invalid input: try putting a number there")

    while budget_information[4] < 0.0:
        num_input = input("What is your current mortgage or rent per month? ")
        try:
            budget_information[4] = float(num_input)
            if budget_information[4] < 0.0:
                print("Invalid input: mortgage or rent cannot be negative.")
        except:
            print("Invalid input: try putting a number there")
    
    while budget_information[5] < 0.0:
        num_input = input("How much do you pay for your car loans per month? ")
        try:
            budget_information[5] = float(num_input)
            if budget_information[5] < 0.0:
                print("Invalid input: car loan payment cannot be negative.")
        except:
            print("Invalid input: try putting a number there")
    
    while budget_information[6] < 0.0:
        num_input = input("What are your current debt payments besides mortgage and car loan? ")
        try:
            budget_information[6] = float(num_input)
            if budget_information[6] < 0.0:
                print("Invalid input: debt payments cannot be negative.")
        except:
            print("Invalid input: try putting a number there")

    while budget_information[7] < 0.0:
        num_input = input("How much do you pay on utilities every month? ")
        try:
            budget_information[7] = float(num_input)
            if budget_information[7] < 0.0:
                print("Invalid input: utilities cannot be negative.")
        except:
            print("Invalid input: try putting a number there")
    
    while budget_information[8] < 0.0:
        num_input = input("How much do you pay on insurance (medical, pet, auto, home, life) every month? ")
        try:
            budget_information[8] = float(num_input)
            if budget_information[8] < 0.0:
                print("Invalid input: insurance cannot be negative.")
        except:
            print("Invalid input: try putting a number there")
    
    while budget_information[9] <= 0.0:
        num_input = input("How much do you spend on food each month. Include doordash, grocery store, and restaurants.")
        try:
            budget_information[9] = float(num_input)
            if budget_information[9] <= 0.0:
                print("Invalid input: You must spend money on food.")
        except:
            print("Invalid input: try putting a number there")
    
    while budget_information[10] < 0.0:
        num_input = input("Do you have any medical co-pays?")
        try:
            budget_information[10] = float(num_input)
            if budget_information[10] < 0.0:
                print("Invalid input: medical copays cannot be negative.")
        except:
            print("Invalid input: try putting a number there")
    
    while budget_information[11] < 0.0:
        num_input = input("How much money do you give every month? ")
        try:
            budget_information[11] = float(num_input)
            if budget_information[11] < 0.0:
                print("Invalid input: giving cannot be negative.")
        except:
            print("Invalid input: try putting a number there")
    
    while budget_information[12] < 0.0:
        num_input = input("How much money do you invest every month? ")
        try:
            budget_information[12] = float(num_input)
            if budget_information[12] < 0.0:
                print("Invalid input: investments cannot be negative.")
        except:
            print("Invalid input: try putting a number there")
    
    while budget_information[13] < 0.0:
        num_input = input("How much money do you spend on extra expenditures beyond those listed here? ")
        try:
            budget_information[13] = float(num_input)
            if budget_information[13] < 0.0:
                print("Invalid input: extra expenditures cannot be negative.")
        except:
            print("Invalid input: try putting a number there")

    generate_budget_document(budget_types, budget_information, fileExists)
    return fileExists

#def generate_budget_document(types, info, fileExists = False):

    workbook = None
    worksheet = None

    giving_ten_percent = info[0] / 10
    food_cost_min = info[1] * 250
    extras_cost_min = info[1] * 20 + 60
    pet_cost_min = info[2] * 20
    pet_insurance_min = info[2] * 20
    insurance_min = info[3] * 50 + 150
    
    if(not fileExists):
        workbook = Workbook()
        worksheet = workbook.active
        fileExists = True
    else:
        workbook = openpyxl.load_workbook("InterestCalculation.xlsx")
        worksheet = workbook.create_sheet()
    
    worksheet.title = "Budget creator $%.2f" % info[0]

    worksheet["A1"] = "Budget calculation"
    worksheet["A2"] = "Monthly pay: $%.2f" % info[0]
    worksheet["C1"] = "Current Spending"
    worksheet["D1"] = "Frugal Chart"
    worksheet["E1"] = "Generous Chart"
    money_format = "$#,##0.00"

    for i in range(len(types)):
        worksheet["B%d" % (i + 2)] = types[i]
    
    worksheet["B16"] = "Total expenditures"
    worksheet["B17"] = "Remaining Funds"

    for c in "CDE":
        worksheet[c + "2"] = info[0]
        worksheet[c + "2"].number_format = money_format
        worksheet[c + "3"] = info[1]
        worksheet[c + "4"] = info[2]
        worksheet[c + "5"] = info[3]
        worksheet[c + "6"] = info[4]
        worksheet[c + "7"] = info[5]
        worksheet[c + "8"] = info[6]
        worksheet[c + "9"] = info[7]
        worksheet[c + "12"] = info[10]
    
    worksheet["C10"] = info[8]
    worksheet["C11"] = info[9]
    worksheet["C13"] = info[11]
    worksheet["C14"] = info[12]
    worksheet["C15"] = info[13]
    worksheet["C16"] = f"=SUM(C6:C15)"
    worksheet["C17"] = f"=C2-C16"
    worksheet["D10"] = insurance_min + pet_insurance_min
    worksheet["D11"] = food_cost_min
    worksheet["D13"] = min(100, giving_ten_percent)
    worksheet["D14"] = 250
    worksheet["D15"] = extras_cost_min + pet_cost_min
    worksheet["D16"] = f"=SUM(D6:D15)"
    worksheet["D17"] = f"=D2-D16"
    worksheet["E10"] = (insurance_min + pet_insurance_min) * 1.25
    worksheet["E11"] = food_cost_min * 1.5
    worksheet["E13"] = giving_ten_percent
    worksheet["E14"] = 1000
    worksheet["E15"] = extras_cost_min + pet_cost_min + 250
    worksheet["E16"] = f"=SUM(E6:E15)"
    worksheet["E17"] = f"=E2-E16"

    for c in "CDE":
        for i in range(6, 18):
            worksheet["%c%d" % (c, i)].number_format = money_format

    for col in worksheet.columns:
        length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if(len(str(cell.value)) > length):
                    length = len(str(cell.value))
            except:
                pass
        worksheet.column_dimensions[column].width = length + 2
    
    generate_pie_chart(worksheet, 2, 3)
    generate_pie_chart(worksheet, 2, 4)
    generate_pie_chart(worksheet, 2, 5)
    workbook.save("InterestCalculation.xlsx")
    print("Finished creating budget documents.")

        
    return fileExists

#def generate_pie_chart(worksheet, label_col, data_col):
    chart = PieChart()
    labels = Reference(worksheet, min_col=label_col, min_row=6, max_row=15)
    data = Reference(worksheet, min_col=data_col, min_row=5, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    if data_col == 3:
        chart.title = "Current Budget Makeup"
        worksheet.add_chart(chart, "H3")
    elif data_col == 4:
        chart.title = "Frugal Budget Makeup"
        worksheet.add_chart(chart, "H19")
    elif data_col == 5:
        chart.title = "Generous Budget Makeup"
        worksheet.add_chart(chart, "Q19")
    return