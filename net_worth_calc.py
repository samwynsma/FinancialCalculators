import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook

class NetWorthCalculator:
    def __init__(self, file_exists = False):
        self.file_exists = file_exists
        self.excel_file = "InterestCalculation.xlsx"
        self.categories = []
        self.amount_per_category = []
        self.total = 0.0
        self.percentile = 0.0
        self.window = None
    
    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def create_gui(self):
        window = tk.Toplevel()
        self.window = window
        window.title("Net Worth Calculator")
        window.geometry("460x520")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Net Worth Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Type in an item and an amount, indicate whether it is an asset (something you own) or a liability (debt). Calculate when you have everything.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))

        net_frame = tk.Frame(window)
        net_frame.pack(padx=20, pady=8, fill="x")

        tk.Label(net_frame, text="Item:", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.item_entry = tk.Entry(net_frame, width=28)
        self.item_entry.grid(row=0, column=1, pady=6)

        tk.Label(net_frame, text="Value:", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.worth_entry = tk.Entry(net_frame, width=28)
        self.worth_entry.grid(row=1, column=1, pady=6)

        tk.Button(net_frame, text="Asset", width=16, command=self.add_asset).grid(row=0, column=2, padx=6)
        tk.Button(net_frame, text="Liability", width=16, command=self.add_liability).grid(row=1, column=2, padx=6)

        values_frame = tk.LabelFrame(window, text="Assets and Liabilities", padx=10, pady=10)
        values_frame.pack(padx=20, pady=(8, 10), fill="both", expand=True)

        results_canvas = tk.Canvas(values_frame, height=140, highlightthickness=0)
        results_scrollbar = tk.Scrollbar(values_frame, orient="vertical", command=results_canvas.yview)
        results_canvas.configure(yscrollcommand=results_scrollbar.set)

        results_inner = tk.Frame(results_canvas)
        results_canvas.create_window((0, 0), window=results_inner, anchor="nw")

        self.results_text = tk.Text(
            results_inner,
            wrap="word",
            height=12,
            yscrollcommand=results_scrollbar.set,
            state="normal",
            padx=6,
            pady=6,
        )
        self.results_text.pack(fill="both", expand=True)

        results_canvas.pack(side="left", fill="both", expand=True)
        results_scrollbar.pack(side="right", fill="y")

        def on_canvas_configure(event):
            results_canvas.configure(scrollregion=results_canvas.bbox("all"))

        results_inner.bind("<Configure>", on_canvas_configure)
        results_canvas.bind("<Configure>", on_canvas_configure)

        self.results_text.configure(state="disabled")

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Calculate", width=16, command=self.on_calculate).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)
        tk.Button(button_frame, text="Percentiles", width=16, command=self.display_ranges).grid(row=0, column=2, padx=6)

        window.grab_set()
        window.mainloop()

    def get_percentile_values(self):
        return [-76472.00, -45428.00, -26450.00, -14983.00, -9878.00, -4381.00, -831.80, 1.00, 182.20, 440.20, 
                990.20, 2552.00, 4056.00, 5208.00, 6532.20, 7726.00, 9256.00, 10370.40, 11810.00, 13528.00,
                15600.20, 18022.20, 20716.00, 23310.00, 27016.00, 30316.20, 34242.00, 39436.00, 44734.00, 51366.00,
                57040.00, 62600.20, 67500.00, 73120.20, 79054.00, 84256.00, 89534.00, 96524.00, 101964.00, 110314.00,
                117810.00, 125686.00, 132632.00, 141164.00, 147316.00, 155908.00, 164132.00, 172168.00, 181562.00, 192084.00,
                202106.20, 212562.00, 223554.00, 238034.00, 250380.00, 261644.00, 274944.00, 288614.00, 298844.00, 312622.00,
                327622.00, 347520.00, 366448.00, 384910.00, 402800.00, 415460.00, 429190.00, 447958.00, 468284.20, 493068.00,
                521000.20, 551988.00, 587968.00, 622546.00, 685340.00, 697576.00, 743564.00, 785484.00, 836944.00, 891750.00,
                947453.00, 1009860.00, 1078294.00, 1154634.00, 1234848.00, 1308426.00, 1399334.00, 1510942.00, 1693542.00, 1920758.00,
                2157988.00, 2382960.00, 2692160.00, 3088722.00, 3779600.00, 4699180.20, 6150980.00, 8464740.20, 13666778.00]

    def get_percentile(self):
        percentile_values = self.get_percentile_values()
        for i in range(len(percentile_values)):
            if percentile_values[i] > self.total:
                self.percentile = i
                return
        self.percentile = 99

    def display_ranges(self):
        popup = tk.Toplevel(self.window)
        popup.title("Percentile Ranges")
        popup.geometry("420x420")
        popup.resizable(False, False)

        tk.Label(
            popup,
            text="Net Worth Percentile Thresholds",
            font=("Segoe UI", 12, "bold"),
            pady=8,
        ).pack()

        text_frame = tk.Frame(popup)
        text_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        text_widget = tk.Text(text_frame, wrap="word", height=20, width=48)
        scrollbar = tk.Scrollbar(text_frame, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)

        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        percentile_values = self.get_percentile_values()
        lines = []
        for index, value in enumerate(percentile_values):
            if index == 0:
                lines.append(f"0%: <= {value:,.2f}")
            else:
                previous_value = percentile_values[index - 1]
                lines.append(f"{index}%: {previous_value:,.2f} to {value:,.2f}")
        lines.append(f"99%: >= {percentile_values[len(percentile_values)-1]:,.2f}")

        text_widget.insert("end", "\n".join(lines))
        text_widget.configure(state="disabled")

    def add_asset(self):
        try:
            asset_value = self.parse_float(self.worth_entry.get())
        except:
            messagebox.showerror("Input error", "Please enter valid numeric values for asset value.")
            return
        if(asset_value <= 0.0):
            messagebox.showerror("Input error", "Asset value must be positive")
            return
        asset = self.item_entry.get()
        if(len(asset) == 0):
            messagebox.showerror("Input error", "Please enter an asset.")
            return
        if asset in self.categories:
            messagebox.showerror("Input error", "Asset category already listed.")
            return
        self.categories.append(asset)
        self.amount_per_category.append(asset_value)
        self.results_text.configure(state="normal")
        self.results_text.insert("end", asset + ": " + str(asset_value) + "\n")
        self.results_text.configure(state="disabled")
    
    def add_liability(self):
        try:
            liability_value = self.parse_float(self.worth_entry.get())
        except:
            messagebox.showerror("Input error", "Please enter valid numeric values for liability value.")
            return
        if(liability_value <= 0.0):
            messagebox.showerror("Input error", "Liability value must be positive")
            return
        liability = self.item_entry.get()
        if(len(liability) == 0):
            messagebox.showerror("Input error", "Please enter an liability.")
            return
        if liability in self.categories:
            messagebox.showerror("Input error", "Liability category already listed.")
            return
        self.categories.append(liability)
        self.amount_per_category.append(-liability_value)
        self.results_text.configure(state="normal")
        self.results_text.insert("end", liability + ": " + str(-liability_value) + "\n")
        self.results_text.configure(state="disabled")
    
    def on_calculate(self):
        self.total = 0.0
        for item in self.amount_per_category:
            self.total += item
        self.get_percentile()
        self.file_exists = self.generate_net_worth_document()
        self.result_label.config(text="Your net worth is " + str(self.total) + " which puts you at the " + str(self.percentile) + " percentile of Americans. Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("Inheritance document complete", "Inheritance saved to InterestCalculation.xlsx.")
    
    def generate_net_worth_document(self):
        percentile = self.percentile
        assets = []
        liabilities = []
        asset_vals = []
        liability_vals = []
        total_asset = 0.0
        total_liability = 0.0
        for i in range(len(self.amount_per_category)):
            item_value = self.amount_per_category[i]
            if item_value < 0:
                liabilities.append(self.categories[i])
                liability_vals.append(self.amount_per_category[i])
                total_liability -= self.amount_per_category[i]
            else:
                assets.append(self.categories)
                asset_vals.append(self.amount_per_category[i])
                total_asset += self.amount_per_category[i]
        
        total_net = total_asset - total_liability

        money_format = "$#,##0.00"

        worksheet = None
        workbook = None

        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook(self.excel_file)
            worksheet = workbook.create_sheet()

        worksheet.title = "Net Worth %d" % self.total

        worksheet["A1"] = "Net Worth Calculator"
        worksheet["B2"] = "Assets"
        worksheet["C2"] = "Asset Value"
        worksheet["D2"] = "Liabilities"
        worksheet["E2"] = "Liability Value"
        worksheet["F3"] = "Total Asset Value"
        worksheet["F4"] = "Total Liability Value"
        worksheet["F5"] = "Total Net Worth"
        worksheet["F6"] = "Net Worth Percentile"
        worksheet["G2"] = "Full Information"

        for i in range(len(assets)):
            worksheet["B%d" % (i+2)] = assets[i]
            worksheet["C%d" % (i+2)] = asset_vals[i]

        for i in range(len(liabilities)):
            worksheet["D%d" % (i+2)] = liabilities[i]
            worksheet["E%d" % (i+2)] = liability_vals[i]
        
        worksheet["G3"] = total_asset
        worksheet["G4"] = total_liability
        worksheet["G5"] = total_net
        worksheet["G6"] = self.percentile

        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2
        

        workbook.save(self.excel_file)
        print("Finished creating net worth documents.")
        return self.file_exists

        

def net_worth(file_exists):
    calculator = NetWorthCalculator(file_exists)
    calculator.create_gui()
    return calculator.file_exists