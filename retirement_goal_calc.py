import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook

from excel_doc import ExcelDocument


class RetirementCalculator:
    def __init__(self, file_exists=False, document=None):
        self.document = document if document is not None else ExcelDocument()
        self.excel_file = "InterestCalculation.xlsx"
        self.monthly_needs = -1.0
        self.current_savings = -1.0
        self.growth_rate = -1.0
        self.current_age = -1
        self.retirement_age = -1
        self.death_age = -1
        self.file_exists = file_exists
    
    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def parse_int(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int)):
            return int(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return int(raw_value)
    
    def create_gui(self):
        window = tk.Toplevel()
        window.title("Retirement Goals Calculator")
        window.geometry("460x460")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Retirement Goals Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Enter your retirement goal details below and click Calculate.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))

        inv_frame = tk.Frame(window)
        inv_frame.pack(padx=20, pady=8, fill = "x")

        tk.Label(inv_frame, text="Current Savings ($):", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.savings_entry = tk.Entry(inv_frame, width=28)
        self.savings_entry.grid(row=0, column=1, pady=6)

        tk.Label(inv_frame, text="Annual interest rate (%):", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.interest_rate_entry = tk.Entry(inv_frame, width=28)
        self.interest_rate_entry.grid(row=1, column=1, pady=6)

        tk.Label(inv_frame, text="Monthly Needs ($):", anchor="w").grid(row=2, column=0, sticky="w", pady=6)
        self.needs_entry = tk.Entry(inv_frame, width=28)
        self.needs_entry.grid(row=2, column=1, pady=6)

        tk.Label(inv_frame, text="Current Age:", anchor="w").grid(row=3, column=0, sticky="w", pady=6)
        self.age_entry = tk.Entry(inv_frame, width=28)
        self.age_entry.grid(row=3, column=1, pady=6)

        tk.Label(inv_frame, text="Retirement Age:", anchor="w").grid(row=4, column=0, sticky="w", pady=6)
        self.retirement_entry = tk.Entry(inv_frame, width=28)
        self.retirement_entry.grid(row=4, column=1, pady=6)

        tk.Label(inv_frame, text="Life Expectancy:", anchor="w").grid(row=5, column=0, sticky="w", pady=6)
        self.end_entry = tk.Entry(inv_frame, width=28)
        self.end_entry.grid(row=5, column=1, pady=6)

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Calculate", width=16, command=self.on_calculate).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)

        window.grab_set()
        window.mainloop()
    
    def on_calculate(self):
        try:
            starting_savings = self.parse_float(self.savings_entry.get())
            interest = self.parse_float(self.interest_rate_entry.get())
            needs = self.parse_float(self.needs_entry.get())
            current_age = self.parse_int(self.age_entry.get())
            retirement_age = self.parse_int(self.retirement_entry.get())
            end_age = self.parse_int(self.end_entry.get())
        except ValueError:
            messagebox.showerror("Input error", "Please enter valid numeric values for all fields.")
            return
        
        if(starting_savings <= 0.0):
            messagebox.showerror("Input error", "Starting savings cannot be negative. That's called a loan.")
            return
        
        if(interest <= 0.0):
            messagebox.showerror("Input error", "Interest rate must be positive. This isn't a loan, its an investment.")
            return
        elif(interest > 12.0):
            proceed = messagebox.askyesno(
                "High Interest",
                "Are you sure that your interest is greater than 12 percent? That is an awfully high interest rate.",
            )
            if not proceed:
                return
        
        if(needs <= 0.0):
            messagebox.showerror("Input error", "If you honestly think that you will need no money, you are dumb.")
            return
        
        if(current_age < 18 or current_age >= 70):
            messagebox.showerror("Input error", "Your current age is out of range. Please select an age that is between 18 and 70.")
            return
        
        if(retirement_age < current_age or retirement_age > 70):
            messagebox.showerror("Input error", "Retirement age is out of range. Retirement age must be greater than current age and less than or equal to 70.")
            return
        
        if(end_age <= retirement_age):
            messagebox.showerror("Stupid person error", "Saving for retirement isn't for those who expect to die before they retire.")
            return
        
        if(end_age >= 105):
            messagebox.showerror("Ridiculous age assessment error", "Less than 0.01 percent of people get to that age. Please pick a reasonable death age.")
            return
        
        self.monthly_needs = needs
        self.current_savings = starting_savings
        self.growth_rate = interest
        self.current_age = current_age
        self.retirement_age = retirement_age
        self.death_age = end_age

        self.file_exists = self.generate_retirement_documents()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("Retirement document complete", "Retirement calculation saved to InterestCalculation.xlsx.")
        
    def generate_retirement_documents(self):
        month_age = self.current_age * 12
        month_ret = self.retirement_age * 12
        month_death = self.death_age * 12
        month_growth = self.growth_rate / 12.0
        current_savings = self.current_savings
        monthly_needs = self.monthly_needs

        age_to_ret_gap = month_ret - month_age
        ret_to_death_gap = month_death - month_ret

        scenario_one = current_savings # Scenario one: money lasts until death plus five years
        scenario_two = current_savings # Scenario two: account maintains itself: growth rate = take out rate.
        scenario_three = current_savings # Scenario three: account user takes out half the growth at the starting year.

        scenario_one_goal = monthly_needs * (1 - pow(1 + (month_growth / 100.0), -(ret_to_death_gap + 60))) / (month_growth / 100.0)
        scenario_two_goal = monthly_needs / (month_growth / 100.0)
        scenario_three_goal = monthly_needs / (month_growth / 200.0)

        print(round(scenario_one_goal, 2))
        print(scenario_two_goal)
        print(scenario_three_goal)
        print(" ")

        month_growth_percent = month_growth / 100.0
        numerator_calc = current_savings * pow(1 + month_growth_percent, age_to_ret_gap)
        denom_calc1 = (pow(1 + month_growth_percent, age_to_ret_gap) - 1.0) / month_growth_percent
        scenario_one_monthly = (scenario_one_goal - numerator_calc) / denom_calc1
        scenario_two_monthly = (scenario_two_goal - numerator_calc) / denom_calc1
        scenario_three_monthly = (scenario_three_goal - numerator_calc) / denom_calc1

        print(scenario_one_monthly)
        print(scenario_two_monthly)
        print(scenario_three_monthly)

        money_s1 = []
        money_s2 = []
        money_s3 = []

        money_s1.append(scenario_one)
        money_s2.append(scenario_two)
        money_s3.append(scenario_three)

        for i in range(age_to_ret_gap):
            scenario_one = round(scenario_one * (1 + (month_growth / 100.0)), 2) + scenario_one_monthly
            scenario_two = round(scenario_two * (1 + (month_growth / 100.0)), 2) + scenario_two_monthly
            scenario_three = round(scenario_three * (1 + (month_growth / 100.0)), 2) + scenario_three_monthly
            money_s1.append(scenario_one)
            money_s2.append(scenario_two)
            money_s3.append(scenario_three)
        
        workbook = None
        worksheet = None

        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook(self.excel_file)
            worksheet = workbook.create_sheet()
        

        worksheet.title = "Retirement %d" % int(monthly_needs)

        worksheet["A1"] = "Retirement Calculation"
        worksheet["A2"] = "Current savings: $%.2f" % current_savings
        worksheet["A3"] = "Monthly needs: $%.2f" % monthly_needs
        worksheet["A4"] = "Growth rate: %.4f percent per month" % self.growth_rate
        worksheet["A5"] = "For money to last until death: $%.2f per month." % scenario_one_monthly
        worksheet["A6"] = "For money to be sustainable: $%.2f per month." % scenario_two_monthly
        worksheet["A7"] = "For money to be economy-resilient: $%.2f per month." % scenario_three_monthly
        worksheet["C1"] = "Last until death plus five years for safety."
        worksheet["F1"] = "Last indefinitely, doesn't grow."
        worksheet["I1"] = "Grows at a reasonable rate."
        worksheet["B2"] = "Month"
        worksheet["C2"] = "Money in account"
        worksheet["E2"] = "Month"
        worksheet["F2"] = "Money in Account"
        worksheet["H2"] = "Month"
        worksheet["I2"] = "Money in Account"
        money_format = "$#,##0.00"

        for i in range(len(money_s1)):
            worksheet["B%d" % (i+3)] = i
            worksheet["C%d" % (i+3)] = money_s1[i]
            worksheet["C%d" % (i+3)].number_format = money_format
        
        for i in range(len(money_s2)):
            worksheet["E%d" % (i+3)] = i
            worksheet["F%d" % (i+3)] = money_s2[i]
            worksheet["F%d" % (i+3)].number_format = money_format
        
        for i in range(len(money_s3)):
            worksheet["H%d" % (i+3)] = i
            worksheet["I%d" % (i+3)] = money_s3[i]
            worksheet["I%d" % (i+3)].number_format = money_format

        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2
        

        workbook.save(self.excel_file)
        print("Finished creating retirement documents.")
        return self.file_exists
        


def retirement_goal(document_or_file_exists=None):
    if isinstance(document_or_file_exists, ExcelDocument):
        calculator = RetirementCalculator(document=document_or_file_exists)
    else:
        calculator = RetirementCalculator(file_exists=document_or_file_exists)
    calculator.create_gui()
    return calculator.file_exists