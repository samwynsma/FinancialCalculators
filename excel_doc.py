import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook

from excel_page import ExcelPage

class ExcelDocument:
    def __init__(self):
        self.pages = 0
        self.page_list = []
        self.current_page = 0

    def create_page(self, info):
        page = ExcelPage(info)
        self.page_list.append(page)
        self.pages = len(self.page_list)
        if self.pages == 1:
            self.current_page = 0
        return page

    @property
    def current_page_title(self):
        if not self.page_list:
            return "No page available"

        page = self.page_list[self.current_page]
        if getattr(page, "title", ""):
            return page.title
        return f"Page {self.current_page + 1}"

    def print_document_xls(self):
        workbook = Workbook()
        worksheet = workbook.active

        for page_index, page in enumerate(self.page_list):
            if page_index > 0:
                worksheet = workbook.create_sheet()

            for column_index, column in enumerate(page.columns, start=1):
                for row_index, value in enumerate(column, start=1):
                    worksheet.cell(row=row_index, column=column_index, value=value)
            

        workbook.save("FinancialDocuments.xlsx")
        print("XLS document created.")

    def print_document_word(self):
        return

    def print_document_csv(self):
        return


