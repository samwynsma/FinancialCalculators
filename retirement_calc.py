import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook

class RetirementDurationCalculator:
    def __init__(self, file_exists = False):
        self.current_savings = -1.0
        self.growth_rate = -1.0
        self.social_security = -1.0
        self.pension = -1.0
        self.monthly_needs = -1.0
        self.expected_years_left = -1
        self.file_exists = file_exists
    
    def parse_float(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return float(raw_value)
    
    def parse_int(self, raw_value):
        if raw_value is None:
            raise ValueError("Missing value")
        if isinstance(raw_value, (int)):
            return int(raw_value)
        raw_value = raw_value.strip().replace(",", "")
        if raw_value.endswith("%"):
            raw_value = raw_value[:-1]
        if raw_value.startswith("$"):
            raw_value = raw_value[1:]
        return int(raw_value)
    
    def create_gui(self):
        window = tk.Toplevel()
        window.title("Retirement Duration Calculator")
        window.geometry("460x460")
        window.resizable(False, False)

        header = tk.Label(
            window,
            text="Retirement Duration Calculator",
            font=("Segoe UI", 16, "bold"),
            wraplength=420,
            justify="center",
            pady=12,
        )
        header.pack()

        instructions = tk.Label(
            window,
            text="Enter your current savings and information below to make a retirement plan.",
            font=("Segoe UI", 10),
            wraplength=420,
            justify="center",
        )
        instructions.pack(pady=(0, 10))

        inv_frame = tk.Frame(window)
        inv_frame.pack(padx=20, pady=8, fill = "x")

        tk.Label(inv_frame, text="Current Savings ($):", anchor="w").grid(row=0, column=0, sticky="w", pady=6)
        self.savings_entry = tk.Entry(inv_frame, width=28)
        self.savings_entry.grid(row=0, column=1, pady=6)

        tk.Label(inv_frame, text="Annual interest rate (%):", anchor="w").grid(row=1, column=0, sticky="w", pady=6)
        self.interest_rate_entry = tk.Entry(inv_frame, width=28)
        self.interest_rate_entry.grid(row=1, column=1, pady=6)

        tk.Label(inv_frame, text="Monthly Needs ($):", anchor="w").grid(row=2, column=0, sticky="w", pady=6)
        self.needs_entry = tk.Entry(inv_frame, width=28)
        self.needs_entry.grid(row=2, column=1, pady=6)

        tk.Label(inv_frame, text="Social Security ($):", anchor="w").grid(row=3, column=0, sticky="w", pady=6)
        self.social_security_entry = tk.Entry(inv_frame, width=28)
        self.social_security_entry.grid(row=3, column=1, pady=6)

        tk.Label(inv_frame, text="Pension ($):", anchor="w").grid(row=4, column=0, sticky="w", pady=6)
        self.retirement_entry = tk.Entry(inv_frame, width=28)
        self.retirement_entry.grid(row=4, column=1, pady=6)

        tk.Label(inv_frame, text="Expected Remaining Life Expectancy:", anchor="w").grid(row=5, column=0, sticky="w", pady=6)
        self.meet_Jesus_entry = tk.Entry(inv_frame, width=28)
        self.meet_Jesus_entry.grid(row=5, column=1, pady=6)

        self.result_label = tk.Label(window, text="", font=("Segoe UI", 10), fg="green", wraplength=420, justify="center")
        self.result_label.pack(pady=(8, 0))

        button_frame = tk.Frame(window)
        button_frame.pack(pady=16)

        tk.Button(button_frame, text="Duration", width=16, command=self.calculate_duration).grid(row=0, column=0, padx=6)
        tk.Button(button_frame, text="Quit", width=16, command=window.destroy).grid(row=0, column=1, padx=6)

        window.grab_set()
        window.mainloop()

    def simulate_vals(self, has_needs=True):
        try:
            savings = self.parse_float(self.savings_entry.get())
            interest = self.parse_float(self.interest_rate_entry.get())
            needs = self.parse_float(self.needs_entry.get())
            government_money = self.parse_float(self.social_security_entry.get())
            pension = self.parse_float(self.retirement_entry.get())
            heaven = self.parse_int(self.meet_Jesus_entry.get())
        except ValueError:
            messagebox.showerror("Input error", "Please enter valid numeric values for all fields.")
            return
        
        if savings <= 0.0:
            messagebox.showerror("Input error", "If you have negative retirement savings, then you screwed up life somewhere and I can't help you.")
            return
        
        if interest <= 0.0:
            messagebox.showerror("Input error", "Interest rate must be positive. Investments usually grow, even in retirement.")
            return
        
        if needs <= 0.0 and has_needs:
            messagebox.showerror("Input error", "You will need to spend money in retirement")
            return
        
        if government_money <= 0.0:
            messagebox.showerror("Input error", "Everyone in the USA receives social security in retirement, provided they worked in their life.")
            return
        
        if pension < 0.0:
            messagebox.showerror("Input error", "Pension must be non-negative")
            return
        
        if heaven <= 0:
            messagebox.showerror("Stupid person error", "Unless you can time-travel backwards to kill yourself, this isn't possible.")
            return
        
        self.current_savings = savings
        self.growth_rate = interest
        self.monthly_needs = needs
        self.social_security = government_money
        self.pension = pension
        self.expected_years_left = heaven

    def calculate_duration(self):
        self.simulate_vals()
        self.file_exists = self.generate_duration_document()
        self.result_label.config(text="Results saved to InterestCalculation.xlsx")
        messagebox.showinfo("Retirement document complete", "Retirement calculation saved to InterestCalculation.xlsx.")
    
    def generate_duration_document(self):
        months_remaining = self.expected_years_left * 12
        current_money = self.current_savings
        monthly_interest = self.growth_rate / 12.0
        pension = self.pension
        social = self.social_security
        needs = self.monthly_needs
        end_of_money = months_remaining + 1

        money = []
        money.append(current_money)

        for i in range(months_remaining):
            current_money = max(0.0, (current_money * monthly_interest) + pension + social - needs)
            if(current_money == 0 and i+1 < end_of_money):
                end_of_money = i+1
            money.append(current_money)
        
        workbook = None
        worksheet = None

        if(not self.file_exists):
            workbook = Workbook()
            worksheet = workbook.active
            self.file_exists = True
        else:
            workbook = openpyxl.load_workbook("InterestCalculation.xlsx")
            worksheet = workbook.create_sheet()
        

        worksheet.title = "Duration %d" % int(self.current_savings)

        worksheet["A1"] = "Duration of Retirement Money"
        worksheet["A2"] = "Starting Money: $%.2f" % self.current_savings
        worksheet["A3"] = "Monthly needs: $%.2f" % needs
        worksheet["A4"] = "Growth rate: %.4f percent per month" % self.growth_rate
        worksheet["A5"] = "Social Security: $%.2f" % social
        worksheet["A6"] = "Pension: $%.2f" % pension
        if end_of_money > months_remaining:
            worksheet["A7"] = "Money at end: $%.2f" % current_money
        else:
            worksheet["A7"] = "Money runs out at month %d" % end_of_money 
        worksheet["B1"] = "Month"
        worksheet["C1"] = "Money in Account Remaining"
        money_format = "$#,##0.00"


        for col in worksheet.columns:
            length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if(len(str(cell.value)) > length):
                        length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = length + 2
        

        workbook.save("InterestCalculation.xlsx")
        print("Finished creating retirement documents.")
        return self.file_exists
    
        


def retirement_dur(file_exists):
    calculator = RetirementDurationCalculator(file_exists)
    calculator.create_gui()
    return calculator.file_exists